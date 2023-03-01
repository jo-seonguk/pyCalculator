import openpyxl
from tkinter import *
import tkinter as tk
import tkinter.ttk

'''
엑셀 데이터를 자신이 원하는 설정으로 변경시켜줌
기술자 분야, 경력, 인정비율 등
'''
root = tk.Tk()
root.geometry("640x900")

# 엑셀에서 사용할 열
workTotal = []
labels = []
radios = []
radioInt = []
workCalc = []
workDic = {}

def calc():
    ws = wb[combox.get()]
    """
    공종, 종별, 공사/용역/설계 일치 확인
    -> 라디오 버튼 몇번이 눌려졌는지 확인
    -> % 셀 [12]에 넣기
    -> t3에 참여일수 수식 입력하기 
    -> v3에 (a x 1) + (b x 0.8) 이런식으로 참여일수 계산값 적기
    -> t5에 금액 수식 입력하기 
    -> v5에 (a x 1) + (b x 0.8) 이런식으로 금액 계산값 적기
    """

    ws.cell(10, 26).number_format = 'yyyy-mm-dd'
    ws.cell(10, 27).number_format = '0'
    ws.cell(10, 28).number_format = '0'

    if year.get() == None or year.get() == '': 
        ws.cell(10, 26).value = '2100-12-31'
    else:
        ws.cell(10, 26).value = str(year.get())
        
    if year.get() == None or year2.get() == '': 
        ws.cell(10, 27).value = 200
    else:
        ws.cell(10, 27).value = int(year2.get())
        
    if year.get() == None or price.get() == '': 
        ws.cell(10, 28).value = 0
    else:
        ws.cell(10, 28).value = int(price.get())
    
   
    
    for idx, row in enumerate(workTotal):
        if radioInt[idx].get() == 0:
            workDic[row] = '100%'
        elif radioInt[idx].get() == 1:
            workDic[row] = '80%'
        elif radioInt[idx].get() == 2:
            workDic[row] = '60%'
        elif radioInt[idx].get() == 3:
            workDic[row] = '0%'

    
    for index, calc in enumerate(workCalc):
        
        if calc[1] in workDic:
            if workDic[calc[1]] == '공종(종별) : 공사/용역/설계':
                calc[1] = '평가적용일수'
            elif workDic[calc[1]] == 'None(None) : None':
                calc[1] = "0"
            else:
                calc[1] = workDic[calc[1]]
        else:
            pass


        if calc[1] == '공종(종별) : 공사/용역/설계':
                calc[1] = '평가적용%'
        elif calc[1] == 'None(None) : None':
                calc[1] = "0"
        else:
            pass

        colNum = calc[0][0:3]
        
        if '|' == colNum[-1]:
            colNum = colNum[0:2]
        elif '|' == colNum[-2]:
            colNum = colNum[0]

        
        if colNum != 'Non':
            print(colNum + " : " + calc[1])
            if colNum.isdigit():
                colNum = str(int(colNum)+2)
                ws.cell(int(colNum), 13).number_format = '0%'
                ws.cell(int(colNum), 14).number_format = '0.0'
                ws.cell(int(colNum), 14).value = '=IF(z10-D' + str(colNum) + '<=' + str(ws.cell(10, 27).value) + '*365, IF(J' + str(colNum) + '>=ab10, L' + str(colNum) + '*M' + str(colNum) + ', 0), 0)' #일수 x %
                ws.cell(int(colNum), 15).number_format = '0.0'
                ws.cell(int(colNum), 15).value = '=IF(z10-D' + str(colNum) + '<=' + str(ws.cell(10, 27).value) + '*365, IF(J' + str(colNum) + '>=ab10, J' + str(colNum) + '*M' + str(colNum) + ', 0), 0)' #일수 x %

                if str(ws.cell(int(colNum), 5).value) == combox2.get() or combox2.get() == '모두 선택':  #선택한 전문분야가 맞는지 확인해서 맞을경우 % 넣고 아닌경우 0%
                    ws.cell(int(colNum), 13).value = calc[1]                #% 넣기
                else:
                    ws.cell(int(colNum), 13).value = '0%'
            else:
                pass
        else:
            pass

                
    ws.cell(4, 21).number_format = '0.0'
    ws.cell(4, 23).number_format = '0.0'
    ws.cell(4, 24).number_format = '0.0'
    ws.cell(4, 25).number_format = '0.0'

    ws.cell(7, 21).number_format = '0.0'
    ws.cell(7, 23).number_format = '0.0'
    ws.cell(7, 24).number_format = '0.0'
    ws.cell(7, 25).number_format = '0.0'

    
    ws.cell(10, 22).number_format = '0.0'
    ws.cell(11, 22).number_format = '0.0'
    
    
    ws.cell(3, 21).value = 'total'
    ws.cell(3, 23).value = '100%'
    ws.cell(3, 24).value = '80%'
    ws.cell(3, 25).value = '60%'
    

    ws.cell(4, 21).value = '=SUM(N3:N' + str(len(workCalc)) + ')'  #일수 총합 
    ws.cell(4, 23).value = '=SUMIF(M3:M' + str(len(workCalc)) + ', "100%", N3:N' + str(len(workCalc)) + ')' #100% 합
    ws.cell(4, 24).value = '=SUMIF(M3:M' + str(len(workCalc)) + ', "80%", N3:N' + str(len(workCalc)) + ')'  #80% 합
    ws.cell(4, 25).value = '=SUMIF(M3:M' + str(len(workCalc)) + ', "60%", N3:N' + str(len(workCalc)) + ')'  #60% 합


    ws.cell(5, 21).value = '=SUM(O3:O' + str(len(workCalc)) + ')'  #액수 총합
    ws.cell(5, 23).value = '=SUMIF(M3:M' + str(len(workCalc)) + ', "100%", O3:O' + str(len(workCalc)) + ')' #100% 합
    ws.cell(5, 24).value = '=SUMIF(M3:M' + str(len(workCalc)) + ', "80%", O3:O' + str(len(workCalc)) + ')'  #80% 합
    ws.cell(5, 25).value = '=SUMIF(M3:M' + str(len(workCalc)) + ', "60%", O3:O' + str(len(workCalc)) + ')'  #60% 합

    wb.save(path + filename)
    
    #계산식(100% 일수x1) + (80% 일수 x0.8) + ... = 총합, 계산식(100% 금액x1) + (80% 금 x0.8) + ... = 총합


    ws.cell(4, 27).value = '="경력 일수 = (" &SUMIF(M3:M' + str(len(workCalc)) + ', "100%", N3:N' + str(len(workCalc)) + \
                           ') &" X 1) + (" &SUMIF(M3:M' + str(len(workCalc)) + ', "80%", N3:N' + str(len(workCalc)) + \
                           ') &" X 0.8) + (" &SUMIF(M3:M' + str(len(workCalc)) + ', "60%", N3:N' + str(len(workCalc)) + \
                           ') &" X 0.6) = "&SUM(N3:N' + str(len(workCalc)) + ')'
    
    ws.cell(5, 27).value = '="금액 (백단위) = (" &SUMIF(M3:M' + str(len(workCalc)) + ', "100%", O3:O' + str(len(workCalc)) + \
                           ') &" X 1) + (" &SUMIF(M3:M' + str(len(workCalc)) + ', "80%", O3:O' + str(len(workCalc)) + \
                           ') &" X 0.8) + (" &SUMIF(M3:M' + str(len(workCalc)) + ', "60%", O3:O' + str(len(workCalc)) + \
                           ') &" X 0.6) = "&SUM(O3:O' + str(len(workCalc)) + ')'
    
    ws.cell(7, 21).value = '365일 기준 ='
    ws.cell(7, 22).value = '=' + str(ws.cell(4, 21).value) + '/365'
    ws.cell(8, 21).value = '360일 기준 ='
    ws.cell(8, 22).value = '=' + str(ws.cell(4, 21).value) + '/360'
    ws.cell(2, 14).value = '평가적용일수'
    ws.cell(2, 15).value = '평가적용금액'

    wb.save(path + filename)
    
def listClear():
    workTotal.clear
    workCalc.clear
    labels.clear
    radios.clear
    radioInt.clear


def typeSelect(event):
    listClear()


    if combox2.get() is None:
        print("제대로 선택")
    else:
        if combox2.get() != "목록 선택":
            ws = wb[combox.get()] 
            notebook.add(frame2, text="% 선택 ")
            notebook.add(frame3, text="% 선택 2")
            notebook.add(frame4, text="% 선택 3")
            notebook.add(frame5, text="계산")

            #result = []
            for row in ws.rows:
                #받아오는 셀의 값이 null인 경우 비어있음 none
                #row[8].value: 1종 2동 등등
                #row[7].value: 공사/용역/설계 
                #row[5].value: 공종(도로, 교량, 하천, 항만)
                #row[4].value: 전문분야: 토질지질, 도로공항

                ws.cell(row[0].row, 13).value = '0%'
                
                #전문분야를 선택했거나 모두선택인지 확인
                if str(row[4].value) == combox2.get() or combox2.get() == '모두 선택':   
                    line = []

                    if str(row[8].value) == '1종' or str(row[8].value) == '2종':
                        s_type = '1,2종'
                    else:
                        s_type = '기타'
                    
                    str1 = str(row[5].value) + '(' + s_type + ') : ' + str(row[7].value)

                    str2 = str(row[0].value) + '|' + str(row[5].value) + '|' + s_type + '|' + str(row[7].value) + '|' + str1
                    line.append(str2)
                    line.append(str1)
                    workCalc.append(line)
                    
                    if row[1].value is None:
                        pass
                    elif row[1].value == "용역명":
                        pass
                    else:
                        if str1 not in workTotal:
                            workTotal.append(str1)
                        else:
                            pass
                else:
                    pass
            
            workTotal.sort()
            #workTotal.reverse()

            for idx, work in enumerate(workTotal):
                radioInt.append(tk.IntVar())
                radios.append([])
                if idx < 33:
                    labels.append(tk.Label(frame2, text=work))
                    radios[idx].append(tk.Radiobutton(frame2, text="100%", variable=radioInt[idx], value="0"))
                    radios[idx].append(tk.Radiobutton(frame2, text="80%", variable=radioInt[idx], value="1"))
                    radios[idx].append(tk.Radiobutton(frame2, text="60%", variable=radioInt[idx], value="2"))
                    radios[idx].append(tk.Radiobutton(frame2, text="0%", variable=radioInt[idx], value="3"))
                    
                elif idx > 66:
                    labels.append(tk.Label(frame4, text=work))
                    radios[idx].append(tk.Radiobutton(frame4, text="100%", variable=radioInt[idx], value="0"))
                    radios[idx].append(tk.Radiobutton(frame4, text="80%", variable=radioInt[idx], value="1"))
                    radios[idx].append(tk.Radiobutton(frame4, text="60%", variable=radioInt[idx], value="2"))
                    radios[idx].append(tk.Radiobutton(frame4, text="0%", variable=radioInt[idx], value="3"))
                else:
                    labels.append(tk.Label(frame3, text=work))
                    
                    radios[idx].append(tk.Radiobutton(frame3, text="100%", variable=radioInt[idx], value="0"))
                    radios[idx].append(tk.Radiobutton(frame3, text="80%", variable=radioInt[idx], value="1"))
                    radios[idx].append(tk.Radiobutton(frame3, text="60%", variable=radioInt[idx], value="2"))
                    radios[idx].append(tk.Radiobutton(frame3, text="0%", variable=radioInt[idx], value="3"))
                    
                labels[idx].grid(row=idx, column=0)
                radios[idx][0].grid(row=idx, column=2, sticky="w")
                radios[idx][1].grid(row=idx, column=3, sticky="w")
                radios[idx][2].grid(row=idx, column=4, sticky="w")
                radios[idx][3].grid(row=idx, column=5, sticky="w")
        else:
            print("선택")

def engineerSelect(event):
    listClear()
    
    if combox.get() is None:
        print("제대로 선택")
    else:
        if combox.get() != "목록 선택":
            ws = wb[combox.get()] 
            #ws = wb.get_sheet_by_name(combox.get())  #오류가 발생하므로 위에 코드를 추천

            for row in ws.rows:
                if str(row[4].value) not in combo_list:

                    if str(row[4].value) == 'None' or str(row[4].value) == '전문분야':
                        pass
                    else:
                        combo_list.append(str(row[4].value))
                else:
                    pass
            combo_list.append('모두 선택')
            combox2['values'] = combo_list
        else:
            print("선택")


# 파일 경로 및 파일 이름(확장자)
path = "C:/Users/User/Documents/test/"
filename = "engineer career.xlsx"

# 엑셀파일 읽어옴 
wb = openpyxl.load_workbook(path + filename)
ws = wb.active

#ArithmeticError

sheet_names = wb.sheetnames
combo_list = []

notebook = tk.ttk.Notebook(root, width=640, height=900)
notebook.pack()

frame1 = tk.Frame(root)
notebook.add(frame1, text="기술자 선택")

frame2 = tk.Frame(root) #%
frame3 = tk.Frame(root) #%2
frame4 = tk.Frame(root) #%3
frame5 = tk.Frame(root) #계산

label_name = tk.Label(frame1, text="기술자 선택")
label_name.grid(row=0, column=0)


combox = tk.ttk.Combobox(frame1, height=len(sheet_names), values=sheet_names)
combox.grid(row=0, column=1)
combox.set("목록 선택")
combox.bind('<<ComboboxSelected>>', engineerSelect)

label_type = tk.Label(frame1, text="전문분야 선택")
label_type.grid(row=1, column=0)

combox2 = tk.ttk.Combobox(frame1, height=len(combo_list), values=combo_list)
combox2.grid(row=1, column=1)
combox2.set("목록 선택")
combox2.bind('<<ComboboxSelected>>', typeSelect)



yearLabel = tk.Label(frame5, text="기준일 :")
yearLabel.grid(row=0, column=0)

year2Label = tk.Label(frame5, text="적용년수 :")
year2Label.grid(row=1, column=0)

priceLabel = tk.Label(frame5, text="최소 실적금액 :")
priceLabel.grid(row=2, column=0)

year = StringVar()
year2 = StringVar()
price = IntVar()

yearEntry = tk.Entry(frame5, textvariable=year)
yearEntry.grid(row=0, column=1)

year2Entry = tk.Entry(frame5, textvariable=year2)
year2Entry.grid(row=1, column=1)

priceEntry = tk.Entry(frame5, textvariable=price)
priceEntry.grid(row=2, column=1)



button = tk.Button(frame5, command=calc, text="계산")
button.grid(row=3, column=1)


root.mainloop()
