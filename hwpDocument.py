import os
import openpyxl
import win32com.client as win32

path = '//DESKTOP-R4LHHT4/test/'
dic = {}
field1_list= ['field1_name', 'field1_birth', 'field1_position', 'field1_grade']
field2_list = ['field2_name', 'field2_code', 'field2_day']
field3_list = ['field3_name', 'field3_code', 'field3_day']
field1 = []
field2 = []
field4 = []
field5 = []

# 한/글 열기

hwp = win32.gencache.EnsureDispatch("hwpframe.hwpobject")
hwp.XHwpWindows.Item(0).Visible = True
hwp.RegisterModule("FilePathCheckDLL", "FilePathCheckerModule")
hwp.Open(r"//DESKTOP-R4LHHT4/test/계약서류총괄(지자체)" + ".hwp")


# 엑셀 열기

wb = openpyxl.load_workbook(path + "계약_총괄.xlsx", data_only=True)
ws= wb.active
ws = wb["밀양시"]


for row in ws.rows:
    dic[str(row[0].value)] = str(row[1].value)


ws2 = wb['참여기술자']

for row in ws2.rows:
    line = []

    for i in range(0, 4):
        if str(row[i].value) == '0':
            break
        else:
            line.append(str(row[i].value))
    if len(line) == 0:
        pass
    else:
        field1.append(line)

    


    line2 = []
    if str(row[0].value) == '0':
        pass
    else:
        line2.append(str(row[0].value))
        line2.append(str(row[4].value))
        line2.append(str(row[5].value))
    if len(line2) == 0:
        pass
    else:
        field2.append(line2)

field4.append(str(ws2.cell(1,1).value))
field4.append(str(ws2.cell(1,3).value))
field4.append(str(ws2.cell(1,9).value))

field5.append(str(ws2.cell(1,1).value))
field5.append(str(ws2.cell(1,9).value))
field5.append(str(ws2.cell(1,5).value))
field5.append(str(ws2.cell(1,7).value))
field5.append(str(ws2.cell(1,4).value))
field5.append(str(ws2.cell(1,8).value))
field5.append(str(ws2.cell(1,6).value))
    
        
#만약 이상한 값이 들어간다면 이 부분을 바꿔야할 가능성 100000%
hwp_list = hwp.GetFieldList(1).split("")  


field1_hwp = []
field2_hwp = []
field3_hwp = []



for i, s in enumerate(hwp_list):
    hwp_str = s[0:-5]
    if hwp_str[-1] == '{':
        hwp_str = s[0:-6]
    else:
        pass

    if hwp_str in dic:
        hwp.PutFieldText(s, dic[hwp_str])

    elif hwp_str in field1_list:
        field1_hwp.append(s)
        
    elif hwp_str in field2_list:
        field2_hwp.append(s)
        
    elif hwp_str in field3_list:
        field3_hwp.append(s)
        
    else:
        pass

for i, s in enumerate(field1_hwp): #용역참가자 편성표 |  이름, 생년월일, 직위, 등급
    if len(field1)*4 == i:
        break
    else:
        hwp.PutFieldText(s, field1[i//4][i%4])

for i, s in enumerate(field2_hwp): #재직증명서 |  이름, 주민, 재직기간
    if len(field2)*3 == i:
        break
    else:
        hwp.PutFieldText(s, field2[i//3][i%3])

for i, s in enumerate(field3_hwp): #보안서약서 |  이름, 주민, 재직기간
    if len(field2)*3 == i:
        break
    else:
        hwp.PutFieldText(s, field2[i//3][i%3])


#현장대리인 신고서
hwp.PutFieldText("member", str(len(field1)))
hwp.PutFieldText('field4_name', field4[0])
hwp.PutFieldText('field4_position', field4[1])
hwp.PutFieldText('field4_birth2', field4[2])


#현장대리인 재직증명서
hwp.PutFieldText('field5_name', field5[0])
hwp.PutFieldText('field5_birth2', field5[1])
hwp.PutFieldText('field5_code', field5[2])
hwp.PutFieldText('field5_address', field5[3])
hwp.PutFieldText('field5_grade', field5[4])
hwp.PutFieldText('field5_day2', field5[5])
hwp.PutFieldText('field5_day', field5[6])

